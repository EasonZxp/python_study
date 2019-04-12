# encoding: utf-8

import os
import time
import signal
import openpyxl
from openpyxl.styles import Font, colors
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

count = 0
tmp_time = 0
is_sigint_up = False
result = []

def get_dev_serialno():
    """function to get DUT's adb serialno"""
    out = os.popen("adb devices").read()
    if (out.split("\n")[1].find("device") == -1):
        print "Pls ensure that the DUT was connectted!"
        return None 
    else:
        return out.split("\n")[1].split("\t")[0]

def dump_logcat(count):
    """function to dump every reboo count logcat log to logcat.txt"""
    os.system("echo =========count: {cnt} LogBegin==============".format(cnt=count) + " >> logcat.txt")
    os.system("adb shell logcat -d -v time >>  logcat.txt")
    os.system("echo =========count: {cnt} LogEnd================ ".format(cnt=count) + " >> logcat.txt")
    return

def openpyxl_generate_xlsx(filename, header):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    #Add header
    ws.append(header)
    #alignment = Alignment(horizontal='right', vertical='right')
    alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws['A1'].alignment = alignment
    ws['B1'].alignment = alignment
    wb.save(filename)
    return

def adjust_sheet_width(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]

    for col in ws.columns:
        max_length = 0
        column_number = col[0].column # Get the column name
        print column_number
        column_letter = get_column_letter(column_number)
        print column_letter
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if (len(str(cell.value)) > max_length):
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        print adjusted_width
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(filename)

def openpyxl_append_xlsx(filename, rows_list):
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]
    #for row in ws.values:
    #    for value in row:
    #        print(value)
    for row in rows_list:
        ws.append(row)
    wb.save(filename)
    return

def sigint_handler(signum, frame):
    global is_sigint_up
    is_sigint_up = True
    print 'catched interrupt signal, please wait test done...'
    return

def loop_reboot_test():
    """function to do reboot test"""
    global count, is_sigint_up, result, tmp_time
    tmp_time = time.clock()
    serialno = get_dev_serialno()
    if serialno == None:
        return False

    signal.signal(signal.SIGINT, sigint_handler)
    #signal.signal(signal.SIGHUP, sigint_handler)
    signal.signal(signal.SIGTERM, sigint_handler)

    openpyxl_generate_xlsx("Test.xlsx", ['Number', 'Time(Seconds)'])
    while True:
        os.system("adb -s " + serialno + " wait-for-device")
        is_boot_completed = os.popen("adb -s " + serialno + " shell getprop sys.boot_completed").read()
        #print "is_boot_completed : %s" % is_boot_completed

        if(is_boot_completed == "1\n"):
            end_time = time.clock()
            deltatime =  end_time - tmp_time
            tmp_time = end_time 
            
            if (count >= 1):
                dump_logcat(count)
                result.append([count, deltatime])

            #write test result to result list
            print [count, deltatime]
            
            count += 1
            print "Do reboot test, count:%d..." % count
            os.system("adb -s " + serialno + " reboot")
        else:
            #print "Not boot completed"
            time.sleep(1)

        if is_sigint_up:
            print "Test done and exit..."
            openpyxl_append_xlsx("Test.xlsx", result)
            adjust_sheet_width("Test.xlsx")
            return True 

loop_reboot_test()
